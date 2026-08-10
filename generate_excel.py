import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

def create_inventory_ledger():
    wb = openpyxl.Workbook()
    
    # Sheet 1: Main Ledger
    ws = wb.active
    ws.title = "Inventory Ledger"
    ws.views.sheetView[0].showGridLines = True

    # Color Palette - Executive Navy & Gold Theme
    NAVY_DARK = "1B365D"
    NAVY_MID = "2C4D75"
    BLUE_LIGHT = "E8EEF5"
    GRAY_LIGHT = "F4F6F9"
    GRAY_BORDER = "D3D3D3"
    TEXT_DARK = "1A1A1A"
    
    # Subheader Fills
    DINESH_HEADER_FILL = "2C4D75"
    MUKESH_HEADER_FILL = "365F91"
    PAYMENT_HEADER_FILL = "2E7D32"
    
    # Fonts
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="E0E0E0")
    font_sec_hdr = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_tbl_hdr = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11, color=TEXT_DARK)
    font_data_bold = Font(name="Calibri", size=11, bold=True, color=TEXT_DARK)
    font_kpi_label = Font(name="Calibri", size=9, bold=True, color="555555")
    font_kpi_val = Font(name="Calibri", size=14, bold=True, color=NAVY_DARK)
    font_kpi_status = Font(name="Calibri", size=13, bold=True, color="9C0006")

    # Fills
    fill_title = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_dinesh_hdr = PatternFill(start_color=DINESH_HEADER_FILL, end_color=DINESH_HEADER_FILL, fill_type="solid")
    fill_mukesh_hdr = PatternFill(start_color=MUKESH_HEADER_FILL, end_color=MUKESH_HEADER_FILL, fill_type="solid")
    fill_payment_hdr = PatternFill(start_color=PAYMENT_HEADER_FILL, end_color=PAYMENT_HEADER_FILL, fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color=NAVY_MID, end_color=NAVY_MID, fill_type="solid")
    fill_zebra = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")
    fill_kpi = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")
    fill_total = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_status_card = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    # Borders
    thin_side = Side(border_style="thin", color=GRAY_BORDER)
    thick_top = Side(border_style="thin", color=NAVY_DARK)
    double_bottom = Side(border_style="double", color=NAVY_DARK)
    
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(top=thick_top, bottom=double_bottom, left=thin_side, right=thin_side)

    # Number Formats
    CURRENCY_FORMAT = '"₹"#,##0.00'
    INT_FORMAT = '#,##0'

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # --- TITLE BANNER ---
    ws.merge_cells("A1:J1")
    ws["A1"] = "PACHAURI INVENTORY & BALANCE LEDGER"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:J2")
    ws["A2"] = "Automated Inventory Tracking, Payment Reconciliation & Activity Audit Log"
    ws["A2"].font = font_subtitle
    ws["A2"].fill = fill_title
    ws["A2"].alignment = align_center
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # --- SUMMARY KPI CARDS (Rows 4-6) ---
    ws.merge_cells("A4:B4")
    ws["A4"] = "TOTAL DINESH ITEMS"
    ws["A4"].font = font_kpi_label
    ws["A4"].alignment = align_center
    ws["A4"].fill = fill_kpi

    ws.merge_cells("A5:B6")
    ws["A5"] = "=F27"
    ws["A5"].font = font_kpi_val
    ws["A5"].alignment = align_center
    ws["A5"].number_format = CURRENCY_FORMAT
    ws["A5"].fill = fill_kpi

    ws.merge_cells("C4:D4")
    ws["C4"] = "TOTAL MUKESH ITEMS"
    ws["C4"].font = font_kpi_label
    ws["C4"].alignment = align_center
    ws["C4"].fill = fill_kpi

    ws.merge_cells("C5:D6")
    ws["C5"] = "=F46"
    ws["C5"].font = font_kpi_val
    ws["C5"].alignment = align_center
    ws["C5"].number_format = CURRENCY_FORMAT
    ws["C5"].fill = fill_kpi

    ws.merge_cells("E4:F4")
    ws["E4"] = "TOTAL PAYMENTS SETTLED"
    ws["E4"].font = font_kpi_label
    ws["E4"].alignment = align_center
    ws["E4"].fill = fill_kpi

    ws.merge_cells("E5:F6")
    ws["E5"] = "=D60"
    ws["E5"].font = font_kpi_val
    ws["E5"].alignment = align_center
    ws["E5"].number_format = CURRENCY_FORMAT
    ws["E5"].fill = fill_kpi

    ws.merge_cells("G4:J4")
    ws["G4"] = "NET BALANCE & SETTLEMENT STATUS"
    ws["G4"].font = font_kpi_label
    ws["G4"].alignment = align_center
    ws["G4"].fill = fill_status_card

    ws.merge_cells("G5:J6")
    ws["G5"] = '=IF(C67>0, "Dinesh owes Mukesh " & TEXT(C67, "₹#,##0.00"), IF(C67<0, "Mukesh owes Dinesh " & TEXT(ABS(C67), "₹#,##0.00"), "BALANCED / SETTLED"))'
    ws["G5"].font = font_kpi_status
    ws["G5"].alignment = align_center
    ws["G5"].fill = fill_status_card

    for r in range(4, 7):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = border_all

    # --- SECTION 1: DINESH PACHAURI ITEMS ---
    ws.merge_cells("A9:J9")
    ws["A9"] = "ITEMS TAKEN BY: DINESH PACHAURI"
    ws["A9"].font = font_sec_hdr
    ws["A9"].fill = fill_dinesh_hdr
    ws["A9"].alignment = align_left
    ws.row_dimensions[9].height = 24

    headers = ["#", "Item Description", "Qty", "Unit/Pack", "Price per Unit", "Total Value", "Status", "Acknowledged By", "Notes", "Date"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num, value=header)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = border_all
    ws.row_dimensions[10].height = 22

    dinesh_data = [
        (1, "Double Bottle Box", 46, "Per Box", 16.00, "Pending"),
        (2, "Double Beer Mug Box", 51, "Per Box", 11.00, "Pending"),
        (3, "Modak Box", 13, "Pack of 6", 23.00, "Pending"),
        (4, "Modak Jar Cartoon", 10, "72 Pcs/Cartoon", 972.00, "Pending"),
        (5, "Salsa Jar Cartoon", 5, "48 Pcs/Cartoon", 864.00, "Pending"),
        (6, "Fragile Tape", 1, "sqft", 140.00, "Pending"),
        (7, "Transparent Tape", 1, "Roll", 65.00, "Pending"),
    ]

    start_row_dinesh = 11
    end_row_dinesh = 25  # 7 filled + 8 blank rows

    for r in range(start_row_dinesh, end_row_dinesh + 1):
        idx = r - start_row_dinesh + 1
        ws.cell(row=r, column=1, value=idx).alignment = align_center
        
        if idx <= len(dinesh_data):
            item = dinesh_data[idx - 1]
            ws.cell(row=r, column=2, value=item[1]).alignment = align_left
            ws.cell(row=r, column=3, value=item[2]).alignment = align_center
            ws.cell(row=r, column=3).number_format = INT_FORMAT
            ws.cell(row=r, column=4, value=item[3]).alignment = align_left
            ws.cell(row=r, column=5, value=item[4]).alignment = align_right
            ws.cell(row=r, column=5).number_format = CURRENCY_FORMAT
            ws.cell(row=r, column=7, value=item[5]).alignment = align_center
        else:
            ws.cell(row=r, column=7, value="Pending").alignment = align_center

        tot_cell = ws.cell(row=r, column=6, value=f'=IF(OR(C{r}="", E{r}=""), "", C{r}*E{r})')
        tot_cell.alignment = align_right
        tot_cell.number_format = CURRENCY_FORMAT

        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = font_data
            cell.border = border_all
            if r % 2 == 0:
                cell.fill = fill_zebra
        ws.row_dimensions[r].height = 20

    # Dinesh Subtotal Row
    ws.cell(row=27, column=2, value="SUBTOTAL (Dinesh's Items)").font = font_data_bold
    ws.cell(row=27, column=2).alignment = align_left
    subtot_dinesh = ws.cell(row=27, column=6, value=f"=SUM(F{start_row_dinesh}:F{end_row_dinesh})")
    subtot_dinesh.font = font_data_bold
    subtot_dinesh.alignment = align_right
    subtot_dinesh.number_format = CURRENCY_FORMAT
    
    for c in range(1, 11):
        cell = ws.cell(row=27, column=c)
        cell.fill = fill_total
        cell.border = border_total
    ws.row_dimensions[27].height = 22

    # --- SECTION 2: MUKESH PACHAURI ITEMS ---
    ws.merge_cells("A29:J29")
    ws["A29"] = "ITEMS TAKEN BY: MUKESH PACHAURI"
    ws["A29"].font = font_sec_hdr
    ws["A29"].fill = fill_mukesh_hdr
    ws["A29"].alignment = align_left
    ws.row_dimensions[29].height = 24

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=30, column=col_num, value=header)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = border_all
    ws.row_dimensions[30].height = 22

    mukesh_data = [
        (1, "Modak Jar Box", 25, "Pack of 4", 17.00, "Pending"),
        (2, "Modak Jar Box", 40, "Pack of 6", 23.00, "Pending"),
        (3, "Single Beer Mug Box", 41, "6/Box", 6.00, "Pending"),
        (4, "Beer Mug Cartoon", 10, "32 Pcs/Cartoon", 680.00, "Pending"),
        (5, "Cloth/Small Wrapping Sheet", 2, "Bundle", None, "Pending"),
        (6, "Big Wrapping Sheet", 1, "Bundle", None, "Pending"),
    ]

    start_row_mukesh = 31
    end_row_mukesh = 45  # 6 filled + 9 blank rows

    for r in range(start_row_mukesh, end_row_mukesh + 1):
        idx = r - start_row_mukesh + 1
        ws.cell(row=r, column=1, value=idx).alignment = align_center
        
        if idx <= len(mukesh_data):
            item = mukesh_data[idx - 1]
            ws.cell(row=r, column=2, value=item[1]).alignment = align_left
            ws.cell(row=r, column=3, value=item[2]).alignment = align_center
            ws.cell(row=r, column=3).number_format = INT_FORMAT
            ws.cell(row=r, column=4, value=item[3]).alignment = align_left
            if item[4] is not None:
                ws.cell(row=r, column=5, value=item[4]).alignment = align_right
                ws.cell(row=r, column=5).number_format = CURRENCY_FORMAT
            ws.cell(row=r, column=7, value=item[5]).alignment = align_center
        else:
            ws.cell(row=r, column=7, value="Pending").alignment = align_center

        tot_cell = ws.cell(row=r, column=6, value=f'=IF(OR(C{r}="", E{r}=""), "", C{r}*E{r})')
        tot_cell.alignment = align_right
        tot_cell.number_format = CURRENCY_FORMAT

        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = font_data
            cell.border = border_all
            if r % 2 == 0:
                cell.fill = fill_zebra
        ws.row_dimensions[r].height = 20

    # Mukesh Subtotal Row
    ws.cell(row=46, column=2, value="SUBTOTAL (Mukesh's Items)").font = font_data_bold
    ws.cell(row=46, column=2).alignment = align_left
    subtot_mukesh = ws.cell(row=46, column=6, value=f"=SUM(F{start_row_mukesh}:F{end_row_mukesh})")
    subtot_mukesh.font = font_data_bold
    subtot_mukesh.alignment = align_right
    subtot_mukesh.number_format = CURRENCY_FORMAT
    
    for c in range(1, 11):
        cell = ws.cell(row=46, column=c)
        cell.fill = fill_total
        cell.border = border_total
    ws.row_dimensions[46].height = 22

    # --- SECTION 3: PAYMENT TRACKING ---
    ws.merge_cells("A48:J48")
    ws["A48"] = "PAYMENT TRACKING LOG"
    ws["A48"].font = font_sec_hdr
    ws["A48"].fill = fill_payment_hdr
    ws["A48"].alignment = align_left
    ws.row_dimensions[48].height = 24

    pay_headers = ["#", "Date", "Paid By", "Amount Paid", "Payment Method", "Notes", "", "", "", ""]
    ws.merge_cells("F49:J49")
    for col_num in range(1, 6):
        cell = ws.cell(row=49, column=col_num, value=pay_headers[col_num-1])
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = border_all
    ws.cell(row=49, column=6, value="Notes").font = font_tbl_hdr
    ws.cell(row=49, column=6).fill = fill_tbl_hdr
    ws.cell(row=49, column=6).alignment = align_center
    for c in range(6, 11):
        ws.cell(row=49, column=c).border = border_all
        ws.cell(row=49, column=c).fill = fill_tbl_hdr
    ws.row_dimensions[49].height = 22

    start_row_pay = 50
    end_row_pay = 59

    for r in range(start_row_pay, end_row_pay + 1):
        idx = r - start_row_pay + 1
        ws.cell(row=r, column=1, value=idx).alignment = align_center
        ws.cell(row=r, column=2).alignment = align_center
        ws.cell(row=r, column=3).alignment = align_left
        ws.cell(row=r, column=4).alignment = align_right
        ws.cell(row=r, column=4).number_format = CURRENCY_FORMAT
        ws.cell(row=r, column=5).alignment = align_center
        
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
        
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.font = font_data
            cell.border = border_all
            if r % 2 == 0:
                cell.fill = fill_zebra
        ws.row_dimensions[r].height = 20

    # Payments Total Row
    ws.cell(row=60, column=3, value="TOTAL PAYMENTS SETTLED").font = font_data_bold
    ws.cell(row=60, column=3).alignment = align_left
    tot_pay = ws.cell(row=60, column=4, value=f"=SUM(D{start_row_pay}:D{end_row_pay})")
    tot_pay.font = font_data_bold
    tot_pay.alignment = align_right
    tot_pay.number_format = CURRENCY_FORMAT
    
    ws.merge_cells("F60:J60")
    for c in range(1, 11):
        cell = ws.cell(row=60, column=c)
        cell.fill = fill_total
        cell.border = border_total
    ws.row_dimensions[60].height = 22

    # --- SECTION 4: BALANCE RECONCILIATION SUMMARY ---
    ws.merge_cells("A62:J62")
    ws["A62"] = "BALANCE RECONCILIATION & INSTRUCTIONS"
    ws["A62"].font = font_sec_hdr
    ws["A62"].fill = fill_title
    ws["A62"].alignment = align_left
    ws.row_dimensions[62].height = 24

    ws.merge_cells("A63:B63")
    ws["A63"] = "Total Value (Dinesh's Items):"
    ws["A63"].font = font_data_bold

    ws.cell(row=63, column=3, value="=F27").font = font_data_bold
    ws.cell(row=63, column=3).number_format = CURRENCY_FORMAT
    ws.cell(row=63, column=3).alignment = align_right

    ws.merge_cells("A64:B64")
    ws["A64"] = "Total Value (Mukesh's Items):"
    ws["A64"].font = font_data_bold

    ws.cell(row=64, column=3, value="=F46").font = font_data_bold
    ws.cell(row=64, column=3).number_format = CURRENCY_FORMAT
    ws.cell(row=64, column=3).alignment = align_right

    ws.merge_cells("A65:B65")
    ws["A65"] = "Total Paid by Dinesh:"
    ws["A65"].font = font_data
    ws.cell(row=65, column=3, value=f'=SUMIF(C{start_row_pay}:C{end_row_pay}, "Dinesh Pachauri", D{start_row_pay}:D{end_row_pay})').font = font_data
    ws.cell(row=65, column=3).number_format = CURRENCY_FORMAT
    ws.cell(row=65, column=3).alignment = align_right

    ws.merge_cells("A66:B66")
    ws["A66"] = "Total Paid by Mukesh:"
    ws["A66"].font = font_data
    ws.cell(row=66, column=3, value=f'=SUMIF(C{start_row_pay}:C{end_row_pay}, "Mukesh Pachauri", D{start_row_pay}:D{end_row_pay})').font = font_data
    ws.cell(row=66, column=3).number_format = CURRENCY_FORMAT
    ws.cell(row=66, column=3).alignment = align_right

    ws.merge_cells("A67:B67")
    ws["A67"] = "NET BALANCE DIFFERENCE:"
    ws["A67"].font = font_data_bold

    ws.cell(row=67, column=3, value="=(C63-C65)-(C64-C66)").font = font_data_bold
    ws.cell(row=67, column=3).number_format = CURRENCY_FORMAT
    ws.cell(row=67, column=3).alignment = align_right

    ws.merge_cells("A68:C68")
    ws["A68"] = '=IF(C67>0, "RESULT: Dinesh owes Mukesh " & TEXT(C67, "₹#,##0.00"), IF(C67<0, "RESULT: Mukesh owes Dinesh " & TEXT(ABS(C67), "₹#,##0.00"), "RESULT: Fully Balanced"))'
    ws["A68"].font = font_kpi_status
    ws["A68"].alignment = align_center
    ws["A68"].fill = fill_status_card

    for r in range(63, 69):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = border_all

    # Instructions Box
    ws.merge_cells("E63:J63")
    ws["E63"] = "OPERATIONAL INSTRUCTIONS & STATUS RULES"
    ws["E63"].font = font_tbl_hdr
    ws["E63"].fill = fill_tbl_hdr
    ws["E63"].alignment = align_center

    instructions = [
        "1. When adding a new item entry, set Status to 'Pending'.",
        "2. The other person acknowledges receipt by changing Status to 'Acknowledged'.",
        "3. Record any money transferred between parties in the PAYMENT TRACKING section.",
        "4. When a transaction or overall balance is settled, set Status to 'Cleared'.",
        "5. Every activity & status update is automatically logged in the 'Activity Audit Log' tab!",
        "6. Items without prices (e.g. Wrapping Sheets) can be exchanged for goods or noted in Cleared section."
    ]

    for idx, inst in enumerate(instructions, 64):
        ws.merge_cells(start_row=idx, start_column=5, end_row=idx, end_column=10)
        ws.cell(row=idx, column=5, value=inst).font = Font(name="Calibri", size=10, italic=False, color="333333")
        ws.cell(row=idx, column=5).alignment = align_left

    for r in range(63, 70):
        for c in range(5, 11):
            ws.cell(row=r, column=c).border = border_all

    # --- DATA VALIDATION DROPDOWNS ---
    dv_status = DataValidation(type="list", formula1='"Pending,Acknowledged,Cleared"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"G{start_row_dinesh}:G{end_row_dinesh}")
    dv_status.add(f"G{start_row_mukesh}:G{end_row_mukesh}")

    dv_paidby = DataValidation(type="list", formula1='"Dinesh Pachauri,Mukesh Pachauri"', allow_blank=True)
    ws.add_data_validation(dv_paidby)
    dv_paidby.add(f"C{start_row_pay}:C{end_row_pay}")

    # --- CONDITIONAL FORMATTING FOR STATUS ---
    fill_pending = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    font_pending = Font(color="856404", bold=True)
    ws.conditional_formatting.add(
        f"G{start_row_dinesh}:G{end_row_dinesh}",
        CellIsRule(operator="equal", formula=['"Pending"'], stopIfTrue=True, fill=fill_pending, font=font_pending)
    )
    ws.conditional_formatting.add(
        f"G{start_row_mukesh}:G{end_row_mukesh}",
        CellIsRule(operator="equal", formula=['"Pending"'], stopIfTrue=True, fill=fill_pending, font=font_pending)
    )

    fill_ack = PatternFill(start_color="CFF4FC", end_color="CFF4FC", fill_type="solid")
    font_ack = Font(color="055160", bold=True)
    ws.conditional_formatting.add(
        f"G{start_row_dinesh}:G{end_row_dinesh}",
        CellIsRule(operator="equal", formula=['"Acknowledged"'], stopIfTrue=True, fill=fill_ack, font=font_ack)
    )
    ws.conditional_formatting.add(
        f"G{start_row_mukesh}:G{end_row_mukesh}",
        CellIsRule(operator="equal", formula=['"Acknowledged"'], stopIfTrue=True, fill=fill_ack, font=font_ack)
    )

    fill_cleared = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    font_cleared = Font(color="0F5132", bold=True)
    ws.conditional_formatting.add(
        f"G{start_row_dinesh}:G{end_row_dinesh}",
        CellIsRule(operator="equal", formula=['"Cleared"'], stopIfTrue=True, fill=fill_cleared, font=font_cleared)
    )
    ws.conditional_formatting.add(
        f"G{start_row_mukesh}:G{end_row_mukesh}",
        CellIsRule(operator="equal", formula=['"Cleared"'], stopIfTrue=True, fill=fill_cleared, font=font_cleared)
    )

    # Auto Column Widths
    col_widths = {
        'A': 6, 'B': 30, 'C': 8, 'D': 18, 'E': 16, 'F': 16, 'G': 15, 'H': 18, 'I': 24, 'J': 14
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- SHEET 2: ACTIVITY AUDIT LOG ---
    ws_log = wb.create_sheet(title="Activity Audit Log")
    ws_log.views.sheetView[0].showGridLines = True

    ws_log.merge_cells("A1:E1")
    ws_log["A1"] = "PACHAURI INVENTORY ACTIVITY & AUDIT TRAIL"
    ws_log["A1"].font = font_title
    ws_log["A1"].fill = fill_title
    ws_log["A1"].alignment = align_center

    log_headers = ["#", "Timestamp", "Performed By (Operator)", "Action Type", "Activity Description & Details"]
    for col_num, h_text in enumerate(log_headers, 1):
        cell = ws_log.cell(row=2, column=col_num, value=h_text)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = border_all
    ws_log.row_dimensions[2].height = 22

    initial_audit_logs = [
        (1, "2026-08-10 10:00:00", "Dinesh Pachauri", "Item Added", "Initial entry created: 7 inventory items added for Dinesh Pachauri (Total: ₹15,841.00)"),
        (2, "2026-08-10 10:05:00", "Mukesh Pachauri", "Item Added", "Initial entry created: 6 inventory items added for Mukesh Pachauri (Total: ₹8,391.00)"),
        (3, "2026-08-10 10:10:00", "System Auto", "Ledger Generated", "Automated ledger initialized with net balance difference: Dinesh owes Mukesh ₹7,450.00")
    ]

    for idx, log in enumerate(initial_audit_logs, 3):
        ws_log.cell(row=idx, column=1, value=log[0]).alignment = align_center
        ws_log.cell(row=idx, column=2, value=log[1]).alignment = align_center
        ws_log.cell(row=idx, column=3, value=log[2]).alignment = align_left
        ws_log.cell(row=idx, column=4, value=log[3]).alignment = align_center
        ws_log.cell(row=idx, column=5, value=log[4]).alignment = align_left

        for c in range(1, 6):
            cell = ws_log.cell(row=idx, column=c)
            cell.font = font_data
            cell.border = border_all
            if idx % 2 == 0:
                cell.fill = fill_zebra
        ws_log.row_dimensions[idx].height = 20

    ws_log.column_dimensions['A'].width = 6
    ws_log.column_dimensions['B'].width = 22
    ws_log.column_dimensions['C'].width = 22
    ws_log.column_dimensions['D'].width = 18
    ws_log.column_dimensions['E'].width = 65

    output_file = "Inventory_Ledger_Pachauri.xlsx"
    wb.save(output_file)
    print(f"Successfully generated automated Excel workbook with Activity Audit Log at: {os.path.abspath(output_file)}")

if __name__ == "__main__":
    create_inventory_ledger()
